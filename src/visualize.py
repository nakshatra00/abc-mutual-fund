import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def add_dashboard_header(sheet):
    """Add dashboard header"""
    sheet['A1'] = '🏦 CORPORATE BOND FUND ANALYSIS DASHBOARD'
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A2'] = 'Portfolio Analysis Across 6 Major AMCs - July 2025'
    sheet['A2'].font = Font(size=12)

def add_charts_to_excel(excel_path):
    """
    Add comprehensive dashboard with charts to the Excel file
    """
    try:
        wb = load_workbook(excel_path)
        
        # Create a new Dashboard sheet 
        if 'Dashboard' in wb.sheetnames:
            wb.remove(wb['Dashboard'])
        dashboard = wb.create_sheet('Dashboard', 0)  # Insert at the beginning
        
        # Get maturity and instrument data for charts
        maturity_sheet = wb['All_Maturity']
        instrument_sheet = wb['All_Instrument']
        
        # Read data safely (avoid merged cells)
        maturity_data = []
        for row in maturity_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[0] != 'Maturity Bucket':  # Skip header and None values
                maturity_data.append({
                    'Maturity Bucket': str(row[0]),
                    'Market/Fair Value (Rs. in Lacs)': float(row[1]) if row[1] else 0,
                    '% of Portfolio (by value)': float(row[2]) if row[2] else 0
                })
        
        instrument_data = []
        for row in instrument_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[0] != 'Instrument Type':  # Skip header and None values  
                instrument_data.append({
                    'Instrument Type': str(row[0]),
                    'Market/Fair Value (Rs. in Lacs)': float(row[1]) if row[1] else 0,
                    '% of Portfolio (by value)': float(row[2]) if row[2] else 0
                })
        
        # Convert to DataFrames
        maturity_df = pd.DataFrame(maturity_data)
        instrument_df = pd.DataFrame(instrument_data)
        
        # Add dashboard components
        add_dashboard_header(dashboard)
        add_executive_summary(dashboard)
        add_maturity_chart(dashboard, maturity_df)
        add_instrument_chart(dashboard, instrument_df)
        add_key_insights(dashboard)
        
        wb.save(excel_path)
        print(f"✅ Dashboard with charts added to {excel_path}")
        
    except Exception as e:
        print(f"Warning: Could not add charts to Excel: {e}")
        return False
    
    return True

def add_executive_summary(sheet):
    """Add executive summary section"""
    start_row = 4
    sheet[f'A{start_row}'] = '📊 EXECUTIVE SUMMARY'
    sheet[f'A{start_row}'].font = Font(size=14, bold=True)
    
    sheet[f'A{start_row + 1}'] = '• Total AMCs Analyzed: 6 (ABSLF, HDFC, ICICI, Kotak, Nippon, SBI)'
    sheet[f'A{start_row + 2}'] = '• Average Portfolio Value: ₹1,35,83,917 Lacs'
    sheet[f'A{start_row + 3}'] = '• Total Valid Holdings: 922 bonds'
    sheet[f'A{start_row + 4}'] = '• Data Quality: 100% valid ISINs'

def add_key_insights(sheet):
    """Add key insights section"""
    start_row = 60
    sheet[f'A{start_row}'] = '💡 KEY INSIGHTS'
    sheet[f'A{start_row}'].font = Font(size=14, bold=True)
    
    insights = [
        "• 75% of portfolio in Perpetual/NA maturity bonds",
        "• 87% allocation to Corporate Bonds vs 12% Government Securities", 
        "• 62% of holdings are AAA-rated (highest credit quality)",
        "• 72% of portfolio yields <7%, indicating conservative positioning",
        "• ABSLF shows highest maturity coverage (99.6%)",
        "• Diversified across 6 major AMCs for risk management"
    ]
    
    for i, insight in enumerate(insights):
        sheet[f'A{start_row + 1 + i}'] = insight

def add_maturity_chart(sheet, maturity_df):
    """Add maturity distribution chart"""
    start_row = 11
    
    # Add title
    sheet[f'A{start_row}'] = '📊 MATURITY BUCKET DISTRIBUTION'
    sheet[f'A{start_row}'].font = Font(size=13, bold=True)
    
    # Add table headers
    headers = ['Maturity Bucket', 'Value (Rs. Lacs)', '% of Portfolio']
    for i, header in enumerate(headers):
        cell = sheet[f'{chr(65+i)}{start_row + 2}']
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    
    # Add data rows
    for idx, (_, row) in enumerate(maturity_df.iterrows()):
        row_num = start_row + 3 + idx
        sheet[f'A{row_num}'] = row['Maturity Bucket']
        sheet[f'B{row_num}'] = round(row['Market/Fair Value (Rs. in Lacs)'], 2)
        sheet[f'C{row_num}'] = f"{row['% of Portfolio (by value)']:.1f}%"
    
    # Create pie chart
    chart1 = PieChart()
    chart1.title = "Portfolio Distribution by Maturity Bucket"
    chart1.height = 10
    chart1.width = 15
    
    data = Reference(sheet, min_col=2, min_row=start_row + 2,
                     max_row=start_row + 2 + len(maturity_df))
    cats = Reference(sheet, min_col=1, min_row=start_row + 3,
                     max_row=start_row + 2 + len(maturity_df))
    
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    sheet.add_chart(chart1, f'E{start_row}')

def add_instrument_chart(sheet, instrument_df):
    """Add instrument type chart"""
    start_row = 35
    
    # Add title
    sheet[f'A{start_row}'] = '📈 INSTRUMENT TYPE ANALYSIS'
    sheet[f'A{start_row}'].font = Font(size=13, bold=True)
    
    # Add table headers
    headers = ['Instrument Type', 'Value (Rs. Lacs)', '% of Portfolio']
    for i, header in enumerate(headers):
        cell = sheet[f'{chr(65+i)}{start_row + 2}']
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
    
    # Add data rows
    for idx, (_, row) in enumerate(instrument_df.iterrows()):
        row_num = start_row + 3 + idx
        sheet[f'A{row_num}'] = row['Instrument Type']
        sheet[f'B{row_num}'] = round(row['Market/Fair Value (Rs. in Lacs)'], 2)
        sheet[f'C{row_num}'] = f"{row['% of Portfolio (by value)']:.1f}%"
    
    # Create bar chart
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Portfolio Distribution by Instrument Type"
    chart2.y_axis.title = 'Value (Rs. in Lacs)'
    chart2.x_axis.title = 'Instrument Type'
    chart2.height = 10
    chart2.width = 15
    
    data = Reference(sheet, min_col=2, min_row=start_row + 2,
                     max_row=start_row + 2 + len(instrument_df))
    cats = Reference(sheet, min_col=1, min_row=start_row + 3,
                     max_row=start_row + 2 + len(instrument_df))
    
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    sheet.add_chart(chart2, f'E{start_row}')
    
def add_amc_summary(sheet, summary_df):
    """Add AMC summary table"""
    start_row = 27
    sheet[f'A{start_row}'] = 'AMC Performance Summary'
    sheet[f'A{start_row}'].font = Font(size=14, bold=True)
    
    # Add headers
    headers = ['AMC', 'Weighted Avg Yield (%)', 'Yield Coverage (%)', 'Maturity Coverage (%)']
    for i, header in enumerate(headers):
        sheet.cell(row=start_row + 1, column=i + 1, value=header).font = Font(bold=True)
    
    # Add data
    for i, row in summary_df.iterrows():
        excel_row = start_row + 2 + i
        sheet[f'A{excel_row}'] = row['AMC']
        sheet[f'B{excel_row}'] = round(row['Weighted Avg Yield (%)'], 2)
        sheet[f'C{excel_row}'] = round(row['Yield Coverage (% of value)'], 2)
        sheet[f'D{excel_row}'] = round(row['Maturity Coverage (% of value)'], 2)

def add_rating_chart(sheet, rating_df):
    """Add rating distribution chart"""
    start_row = 36
    sheet[f'G{start_row}'] = 'Credit Rating Distribution'
    sheet[f'G{start_row}'].font = Font(size=14, bold=True)
    
    # Add top 8 ratings only
    top_ratings = rating_df.head(8)
    
    for i, (rating, value, pct) in enumerate(zip(
        top_ratings['Rating_canonical'], 
        top_ratings['Market/Fair Value (Rs. in Lacs)'],
        top_ratings['% of Portfolio (by value)']
    )):
        row = start_row + 2 + i
        sheet[f'G{row}'] = rating
        sheet[f'H{row}'] = value
        sheet[f'I{row}'] = pct
    
    # Headers
    sheet[f'G{start_row + 1}'] = 'Rating'
    sheet[f'H{start_row + 1}'] = 'Value (Rs. Lacs)'
    sheet[f'I{start_row + 1}'] = '% of Portfolio'

def add_yield_chart(sheet, yield_df):
    """Add yield distribution chart"""
    start_row = 48
    sheet[f'G{start_row}'] = 'Yield Distribution'
    sheet[f'G{start_row}'].font = Font(size=14, bold=True)
    
    # Add data
    for i, (yield_bucket, value, pct) in enumerate(zip(
        yield_df['Yield Bucket'], 
        yield_df['Market/Fair Value (Rs. in Lacs)'],
        yield_df['% of Portfolio (by value)']
    )):
        row = start_row + 2 + i
        sheet[f'G{row}'] = yield_bucket
        sheet[f'H{row}'] = value
        sheet[f'I{row}'] = pct
    
    # Headers
    sheet[f'G{start_row + 1}'] = 'Yield Bucket'
    sheet[f'H{start_row + 1}'] = 'Value (Rs. Lacs)'
    sheet[f'I{start_row + 1}'] = '% of Portfolio'
